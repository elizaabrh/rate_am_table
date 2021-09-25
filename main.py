import re
import requests

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment


def set_up_sheet(sheet):
    # merging some cells
    sheet.merge_cells('D1:E1')
    sheet.merge_cells('F1:G1')
    sheet.merge_cells('H1:I1')
    sheet.merge_cells('J1:K1')

    # setting the font on centre
    i = 1
    j = 4
    while j < 11:
        sheet.cell(i, j).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        j += 2

    # increasing the width of some columns
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15


# this function is for finding the index of a specific character
def find_nth(haystack, needle, n):
    start = haystack.find(needle)
    while start >= 0 and n > 1:
        start = haystack.find(needle, start + len(needle))
        n -= 1
    return start


# this function is for getting all the elements of the table
def get_table(sheet):
    r = requests.get("https://www.rate.am")

    # parsing the text of the page
    page = r.text

    # parsing the html code of the text
    soup = BeautifulSoup(page, 'html.parser')

    # finding all <table> tags with their content
    tables = soup.findChildren('table')

    # getting the table we need
    my_table = tables[3]

    # writing the exchange rates in the first row
    first_row = 1
    j = ord('D')
    for option in my_table.find_all('option'):
        option = str(option)
        if "selected" in option:
            sheet[chr(j) + str(first_row)] = option[option.index(">") + 1:option.index(">") + 6]
            j += 2

    # writing the second row: banks, branches, the date etc
    first_row = 2
    j = ord('A')
    for a in my_table.find_all('a'):
        a = str(a)
        if "Դասակարգել" in a:
            sheet[chr(j) + str(first_row)] = a[a.index('>') + 1:find_nth(a, "<", 2)]
            if "<br/>" in a:
                b = a[a.index('>') + 1:find_nth(a, "<", 2)]
                sheet[chr(j) + str(first_row)] = b + a[find_nth(a, ">", 2) + 1:find_nth(a, "<", 3)]
            j += 1

    # writing the first column: banks
    first_row = 3
    j = ord('A')
    for a in my_table.find_all('a'):
        a = str(a)
        if "անկ" in a or "ԱՆԿ" in a:
            if "Դասակարգել" not in a and "class" not in a:
                sheet[chr(j) + str(first_row)] = a[a.index('>') + 1:find_nth(a, "<", 2)]
                first_row += 1

    # writing the second column: the number of branches
    first_row = 3
    j = ord('B')
    for a_href in my_table.find_all('td'):
        a_href = str(a_href)
        if "a href" in a_href and "bank" in a_href:
            if "class" not in a_href:
                sheet[chr(j) + str(first_row)] = a_href[find_nth(a_href, ">", 2) + 1:find_nth(a_href, "<", 3)]
                first_row += 1

    # writing the third column: the date
    first_row = 3
    j = ord('C')
    for date in my_table.find_all('td'):
        date = str(date)
        if "class=\"date\"" in date:
            sheet[chr(j) + str(first_row)] = date[find_nth(date, ">", 1) + 1:find_nth(date, "<", 2)]
            first_row += 1

    # writing from forth to 11th column: the currency
    data = soup.find_all('td')
    numbers = [d.text for d in data if
               d.text.isdigit() or not len(d.text) or re.match(r'^[-+]?\d+(?:\.\d+)$', d.text)]

    numbers.pop(0)
    first_row = 191
    while first_row >= 153:
        numbers.pop(first_row)
        first_row -= 1

    f = 144
    while f >= 0:
        numbers.pop(f)
        f -= 9

    h = 0
    for first_row in range(3, 20):
        for j in range(ord('D'), ord('L')):
            sheet[chr(j) + str(first_row)] = numbers[h]
            h += 1


if __name__ == "__main__":
    # creating a sheet in excel
    workbook = Workbook()
    sheet = workbook.active
    set_up_sheet(sheet)

    # setting a timer for fetching the data and writing in an excel sheet
    # schedule.every(5).minutes.do(get_table, (sheet, ))

    get_table(sheet)

    workbook.save(filename="data.xlsx")
